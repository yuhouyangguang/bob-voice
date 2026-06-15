import io

from openpyxl import load_workbook

from conftest import auth_headers, login
from app.extensions import db
from app.models import AuditLog, User
from app.services.postprocess import apply_corrections


def test_correction_management_import_export_and_immediate_effect(app, client):
    token = login(client, "admin", "Admin123!")
    headers = auth_headers(token)

    invalid_category = client.post(
        "/api/v1/admin/corrections",
        headers=headers,
        json={"pattern": "错词", "replacement": "正确", "category": "其他"},
    )
    assert invalid_category.status_code == 400

    invalid_regex = client.post(
        "/api/v1/admin/corrections",
        headers=headers,
        json={
            "pattern": "[",
            "replacement": "正确",
            "category": "通用",
            "is_regex": True,
        },
    )
    assert invalid_regex.status_code == 400

    created = client.post(
        "/api/v1/admin/corrections",
        headers=headers,
        json={
            "pattern": "封控",
            "replacement": "风控",
            "category": "风控术语",
        },
    )
    assert created.status_code == 201
    correction = created.get_json()["correction"]
    assert correction["priority"] == len("封控")

    duplicate = client.post(
        "/api/v1/admin/corrections",
        headers=headers,
        json={
            "pattern": "封控",
            "replacement": "风险控制",
            "category": "风控术语",
        },
    )
    assert duplicate.status_code == 400

    with app.app_context():
        corrected, hits = apply_corrections("做好封控工作")
        assert corrected == "做好风控工作"
        assert hits

    csv_content = (
        "pattern,replacement,category,is_regex,enabled,priority\n"
        "封控,风险控制,风控术语,false,true,10\n"
        "北银,北京银行,机构名,false,true,\n"
    ).encode("utf-8")
    imported = client.post(
        "/api/v1/admin/corrections/import",
        headers=headers,
        data={"file": (io.BytesIO(csv_content), "corrections.csv")},
        content_type="multipart/form-data",
    )
    assert imported.status_code == 200
    assert imported.get_json() == {"created": 1, "updated": 1, "total": 2}

    filtered = client.get(
        "/api/v1/admin/corrections",
        headers=headers,
        query_string={
            "category": "风控术语",
            "enabled": "true",
            "sort_by": "priority",
            "order": "desc",
        },
    )
    assert filtered.status_code == 200
    payload = filtered.get_json()
    assert payload["corrections"] == payload["items"]
    assert payload["pagination"]["total"] == 1
    assert payload["items"][0]["replacement"] == "风险控制"
    assert payload["items"][0]["priority"] == 10

    exported = client.get(
        "/api/v1/admin/corrections/export?category=风控术语",
        headers=headers,
    )
    assert exported.status_code == 200
    workbook = load_workbook(io.BytesIO(exported.data), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == (
        "pattern",
        "replacement",
        "category",
        "is_regex",
        "priority",
        "enabled",
    )
    assert rows[1][0:3] == ("封控", "风险控制", "风控术语")

    bad_import = client.post(
        "/api/v1/admin/corrections/import",
        headers=headers,
        data={
            "file": (
                io.BytesIO(
                    (
                        "pattern,replacement,category,is_regex\n"
                        "新词,正确词,通用,false\n"
                        "[,错误正则,通用,true\n"
                    ).encode("utf-8")
                ),
                "bad.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert bad_import.status_code == 400
    assert client.get(
        "/api/v1/admin/corrections?q=新词",
        headers=headers,
    ).get_json()["pagination"]["total"] == 0

    with app.app_context():
        actions = {
            item.action
            for item in AuditLog.query.filter(
                AuditLog.action.in_(
                    {
                        "correction.create",
                        "correction.import",
                        "correction.export",
                    }
                )
            ).all()
        }
        assert actions == {
            "correction.create",
            "correction.import",
            "correction.export",
        }


def test_user_management_crud_permissions_and_password_reset(app, client):
    admin_token = login(client, "admin", "Admin123!")
    headers = auth_headers(admin_token)

    created = client.post(
        "/api/v1/admin/users",
        headers=headers,
        json={
            "username": "advanced001",
            "password": "Start1234",
            "display_name": "高级用户",
            "email": "advanced@example.com",
            "department": "办公室",
            "role": "advanced",
            "is_active": True,
        },
    )
    assert created.status_code == 201
    user = created.get_json()["user"]
    assert user["role"] == "advanced"
    assert "password_hash" not in user

    duplicate = client.post(
        "/api/v1/admin/users",
        headers=headers,
        json={
            "username": "advanced001",
            "password": "Start1234",
            "display_name": "重复用户",
        },
    )
    assert duplicate.status_code == 409

    weak_password = client.post(
        "/api/v1/admin/users",
        headers=headers,
        json={
            "username": "weak001",
            "password": "password",
            "display_name": "弱密码",
        },
    )
    assert weak_password.status_code == 400

    listed = client.get(
        "/api/v1/admin/users",
        headers=headers,
        query_string={"q": "高级", "role": "advanced"},
    )
    assert listed.status_code == 200
    list_payload = listed.get_json()
    assert list_payload["users"] == list_payload["items"]
    assert list_payload["pagination"]["total"] == 1
    assert list_payload["stats"]["admins"] == 1

    updated = client.put(
        f"/api/v1/admin/users/{user['id']}",
        headers=headers,
        json={
            "display_name": "高级用户甲",
            "department": "文秘室",
            "role": "user",
            "is_active": False,
        },
    )
    assert updated.status_code == 200
    assert updated.get_json()["user"]["department"] == "文秘室"
    assert updated.get_json()["user"]["is_active"] is False

    inactive_login = client.post(
        "/api/v1/auth/login",
        json={"username": "advanced001", "password": "Start1234"},
    )
    assert inactive_login.status_code == 401

    reenabled = client.put(
        f"/api/v1/admin/users/{user['id']}",
        headers=headers,
        json={"is_active": True},
    )
    assert reenabled.status_code == 200

    reset = client.post(
        f"/api/v1/admin/users/{user['id']}/reset-password",
        headers=headers,
        json={"password": "Reset5678"},
    )
    assert reset.status_code == 200
    assert login(client, "advanced001", "Reset5678")

    with app.app_context():
        managed_user = db.session.get(User, user["id"])
        managed_user.failed_login_count = 4
        managed_user.locked_until = managed_user.created_at
        db.session.commit()
    unlocked = client.post(
        f"/api/v1/admin/users/{user['id']}/unlock",
        headers=headers,
    )
    assert unlocked.status_code == 200
    assert unlocked.get_json()["user"]["locked_until"] is None

    current_admin = client.get(
        "/api/v1/auth/me",
        headers=headers,
    ).get_json()["user"]
    self_downgrade = client.put(
        f"/api/v1/admin/users/{current_admin['id']}",
        headers=headers,
        json={"role": "user"},
    )
    assert self_downgrade.status_code == 409

    user_token = login(client)
    forbidden = client.post(
        "/api/v1/admin/users",
        headers=auth_headers(user_token),
        json={
            "username": "forbidden001",
            "password": "Password123",
            "display_name": "无权限",
        },
    )
    assert forbidden.status_code == 403

    with app.app_context():
        actions = {
            item.action
            for item in AuditLog.query.filter(
                AuditLog.action.in_(
                    {
                        "user.create",
                        "user.update",
                        "user.password_reset",
                        "user.unlock",
                    }
                )
            ).all()
        }
        assert actions == {
            "user.create",
            "user.update",
            "user.password_reset",
            "user.unlock",
        }
