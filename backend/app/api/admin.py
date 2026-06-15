import csv
import io
import re

from flask import Blueprint, g, jsonify, request, send_file
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_

from ..auth import roles_required
from ..extensions import db
from ..models import Correction, Task, User
from ..utils import audit

bp = Blueprint("admin", __name__, url_prefix="/api/v1")

CORRECTION_CATEGORIES = {
    "产品名",
    "机构名",
    "风控术语",
    "领导表达DNA",
    "通用",
}
CORRECTION_CATEGORY_ALIASES = {
    "finance": "风控术语",
    "person": "领导表达DNA",
    "org": "机构名",
    "product": "产品名",
    "common": "通用",
    "other": "通用",
}
USER_ROLES = {"user", "advanced", "admin"}
IMPORT_MAX_ROWS = 5000
IMPORT_MAX_BYTES = 5 * 1024 * 1024


def _parse_bool(value, field):
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "是", "启用"}:
        return True
    if normalized in {"0", "false", "no", "n", "否", "禁用", ""}:
        return False
    raise ValueError(f"{field} 必须是布尔值")


def _pagination(query):
    page = max(request.args.get("page", 1, type=int), 1)
    per_page = min(max(request.args.get("per_page", 20, type=int), 1), 100)
    result = query.paginate(page=page, per_page=per_page, error_out=False)
    return result, {
        "page": result.page,
        "per_page": result.per_page,
        "total": result.total,
        "pages": result.pages,
    }


def _correction_query():
    query = Correction.query
    search = request.args.get("q", "").strip()
    if search:
        query = query.filter(
            or_(
                Correction.pattern.contains(search),
                Correction.replacement.contains(search),
            )
        )
    category = request.args.get("category", "").strip()
    if category:
        query = query.filter(Correction.category == category)
    enabled = request.args.get("enabled")
    if enabled is not None and enabled != "":
        query = query.filter(Correction.enabled == _parse_bool(enabled, "enabled"))

    sort_by = request.args.get("sort_by", "priority")
    sort_columns = {
        "priority": Correction.priority,
        "pattern": Correction.pattern,
        "category": Correction.category,
        "updated_at": Correction.updated_at,
    }
    column = sort_columns.get(sort_by, Correction.priority)
    order = request.args.get("order", "desc").lower()
    query = query.order_by(
        column.asc() if order == "asc" else column.desc(),
        Correction.id.desc(),
    )
    return query


def _validated_correction(data, existing=None):
    pattern = str(data.get("pattern", existing.pattern if existing else "")).strip()
    replacement = str(
        data.get("replacement", existing.replacement if existing else "")
    ).strip()
    category = str(
        data.get("category", existing.category if existing else "通用")
    ).strip()
    category = CORRECTION_CATEGORY_ALIASES.get(category, category)
    is_regex = _parse_bool(
        data.get("is_regex", existing.is_regex if existing else False),
        "is_regex",
    )
    enabled = _parse_bool(
        data.get("enabled", existing.enabled if existing else True),
        "enabled",
    )
    if not pattern or not replacement:
        raise ValueError("错误模式和正确术语不能为空")
    if len(pattern) > 256 or len(replacement) > 256:
        raise ValueError("错误模式和正确术语最多 256 个字符")
    if category not in CORRECTION_CATEGORIES:
        raise ValueError("category 值无效")
    if is_regex:
        try:
            re.compile(pattern)
        except re.error as exc:
            raise ValueError(f"正则表达式无效: {exc}") from exc
    priority_value = data.get("priority")
    if priority_value in {None, ""}:
        pattern_changed = existing and pattern != existing.pattern
        priority = len(pattern) if not existing or pattern_changed else existing.priority
    else:
        priority = int(priority_value)
    if priority < 0 or priority > 10000:
        raise ValueError("priority 必须在 0-10000 之间")
    duplicate = Correction.query.filter_by(
        pattern=pattern,
        is_regex=is_regex,
    )
    if existing:
        duplicate = duplicate.filter(Correction.id != existing.id)
    if duplicate.first():
        raise ValueError("相同错误模式和规则类型已存在")
    return {
        "pattern": pattern,
        "replacement": replacement,
        "category": category,
        "is_regex": is_regex,
        "priority": priority,
        "enabled": enabled,
    }


@bp.get("/admin/corrections")
@roles_required("admin")
def list_corrections():
    try:
        pagination, meta = _pagination(_correction_query())
    except ValueError as exc:
        return jsonify({"error": "validation_error", "message": str(exc)}), 400
    items = [item.to_dict() for item in pagination.items]
    return {"items": items, "corrections": items, "pagination": meta}


@bp.post("/admin/corrections")
@roles_required("admin")
def create_correction():
    try:
        values = _validated_correction(request.get_json(silent=True) or {})
    except (TypeError, ValueError) as exc:
        return jsonify({"error": "validation_error", "message": str(exc)}), 400
    item = Correction(**values, created_by=g.current_user.id)
    db.session.add(item)
    db.session.flush()
    audit("correction.create", "correction", item.id, values)
    db.session.commit()
    return jsonify({"correction": item.to_dict()}), 201


@bp.put("/admin/corrections/<int:correction_id>")
@roles_required("admin")
def update_correction(correction_id):
    item = db.get_or_404(Correction, correction_id)
    data = request.get_json(silent=True) or {}
    try:
        values = _validated_correction(data, item)
    except (TypeError, ValueError) as exc:
        return jsonify({"error": "validation_error", "message": str(exc)}), 400
    before = item.to_dict()
    for field, value in values.items():
        setattr(item, field, value)
    audit(
        "correction.update",
        "correction",
        item.id,
        {"before": before, "after": values},
    )
    db.session.commit()
    return {"correction": item.to_dict()}


@bp.delete("/admin/corrections/<int:correction_id>")
@roles_required("admin")
def delete_correction(correction_id):
    item = db.get_or_404(Correction, correction_id)
    audit("correction.delete", "correction", item.id, item.to_dict())
    db.session.delete(item)
    db.session.commit()
    return "", 204


def _normalize_import_row(row):
    aliases = {
        "错误模式": "pattern",
        "错误词": "pattern",
        "正确术语": "replacement",
        "替换词": "replacement",
        "分类": "category",
        "是否正则": "is_regex",
        "优先级": "priority",
        "启用": "enabled",
    }
    normalized = {}
    for key, value in row.items():
        key = aliases.get(str(key or "").strip(), str(key or "").strip().lower())
        normalized[key] = value
    return normalized


def _read_correction_rows(file):
    content = file.read(IMPORT_MAX_BYTES + 1)
    if len(content) > IMPORT_MAX_BYTES:
        raise ValueError("导入文件不能超过 5MB")
    suffix = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if suffix == "xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, ())]
        return [dict(zip(headers, values)) for values in rows]
    raise ValueError("仅支持 CSV 或 XLSX 文件")


@bp.post("/admin/corrections/import")
@roles_required("admin")
def import_corrections():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify(
            {"error": "validation_error", "message": "请上传 CSV 或 XLSX 文件"}
        ), 400
    try:
        rows = _read_correction_rows(file)
    except (UnicodeDecodeError, ValueError) as exc:
        return jsonify({"error": "validation_error", "message": str(exc)}), 400
    if len(rows) > IMPORT_MAX_ROWS:
        return jsonify(
            {"error": "validation_error", "message": "单次最多导入 5000 条规则"}
        ), 400

    created = 0
    updated = 0
    errors = []
    for index, raw_row in enumerate(rows, start=2):
        row = _normalize_import_row(raw_row)
        pattern = str(row.get("pattern") or "").strip()
        raw_is_regex = row.get("is_regex", False)
        try:
            is_regex = _parse_bool(raw_is_regex, "is_regex")
            existing = Correction.query.filter_by(
                pattern=pattern,
                is_regex=is_regex,
            ).first()
            values = _validated_correction(row, existing)
            if existing:
                for field, value in values.items():
                    setattr(existing, field, value)
                updated += 1
            else:
                db.session.add(
                    Correction(**values, created_by=g.current_user.id)
                )
                created += 1
        except (TypeError, ValueError) as exc:
            errors.append({"row": index, "message": str(exc)})
    if errors:
        db.session.rollback()
        return jsonify(
            {
                "error": "validation_error",
                "message": "导入文件包含无效数据，未写入任何规则",
                "errors": errors[:100],
            }
        ), 400
    audit(
        "correction.import",
        "correction",
        detail={"created": created, "updated": updated, "filename": file.filename},
    )
    db.session.commit()
    return {"created": created, "updated": updated, "total": created + updated}


@bp.get("/admin/corrections/export")
@roles_required("admin")
def export_corrections():
    try:
        items = _correction_query().all()
    except ValueError as exc:
        return jsonify({"error": "validation_error", "message": str(exc)}), 400
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "术语纠正规则"
    sheet.append(
        ["pattern", "replacement", "category", "is_regex", "priority", "enabled"]
    )
    for item in items:
        sheet.append(
            [
                item.pattern,
                item.replacement,
                item.category,
                item.is_regex,
                item.priority,
                item.enabled,
            ]
        )
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    audit("correction.export", "correction", detail={"count": len(items)})
    db.session.commit()
    return send_file(
        output,
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        as_attachment=True,
        download_name="术语纠正规则.xlsx",
    )


def _validate_password(password):
    if len(password) < 8:
        raise ValueError("密码至少 8 位")
    if not re.search(r"[A-Za-z]", password) or not re.search(r"\d", password):
        raise ValueError("密码必须同时包含字母和数字")


def _active_admin_count():
    return User.query.filter_by(role="admin", is_active=True).count()


def _user_payload(data, existing=None):
    display_name = str(
        data.get("display_name", existing.display_name if existing else "")
    ).strip()
    email = str(data.get("email", existing.email if existing else "") or "").strip()
    department = str(
        data.get("department", existing.department if existing else "") or ""
    ).strip()
    role = str(data.get("role", existing.role if existing else "user")).strip()
    is_active = _parse_bool(
        data.get("is_active", existing.is_active if existing else True),
        "is_active",
    )
    if not display_name:
        raise ValueError("显示姓名不能为空")
    if role not in USER_ROLES:
        raise ValueError("role 值无效")
    return {
        "display_name": display_name[:64],
        "email": email[:128] or None,
        "department": department[:128] or None,
        "role": role,
        "is_active": is_active,
    }


def _serialize_user(user, task_count=0):
    item = user.to_dict()
    item.update(
        {
            "task_count": task_count,
            "last_login": (
                user.last_login_at.isoformat() if user.last_login_at else None
            ),
            "last_login_at": (
                user.last_login_at.isoformat() if user.last_login_at else None
            ),
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "updated_at": user.updated_at.isoformat() if user.updated_at else None,
            "locked_until": (
                user.locked_until.isoformat() if user.locked_until else None
            ),
        }
    )
    return item


@bp.get("/admin/users")
@roles_required("admin")
def list_users():
    task_counts = (
        db.session.query(
            Task.user_id,
            func.count(Task.id).label("task_count"),
        )
        .filter(Task.deleted_at.is_(None))
        .group_by(Task.user_id)
        .subquery()
    )
    query = db.session.query(
        User,
        func.coalesce(task_counts.c.task_count, 0),
    ).outerjoin(task_counts, User.id == task_counts.c.user_id)
    search = request.args.get("q", "").strip()
    if search:
        query = query.filter(
            or_(
                User.username.contains(search),
                User.display_name.contains(search),
                User.department.contains(search),
                User.email.contains(search),
            )
        )
    role = request.args.get("role", "").strip()
    if role:
        query = query.filter(User.role == role)
    active = request.args.get("active")
    try:
        if active is not None and active != "":
            query = query.filter(User.is_active == _parse_bool(active, "active"))
    except ValueError as exc:
        return jsonify({"error": "validation_error", "message": str(exc)}), 400
    query = query.order_by(User.created_at.desc(), User.id.desc())
    pagination, meta = _pagination(query)
    users = [_serialize_user(user, count) for user, count in pagination.items]
    stats = {
        "total": User.query.count(),
        "active": User.query.filter_by(is_active=True).count(),
        "admins": User.query.filter_by(role="admin", is_active=True).count(),
    }
    return {"items": users, "users": users, "pagination": meta, "stats": stats}


@bp.post("/admin/users")
@roles_required("admin")
def create_user():
    data = request.get_json(silent=True) or {}
    username = str(data.get("username", "")).strip()
    password = str(data.get("password", ""))
    if not re.fullmatch(r"[A-Za-z0-9._-]{3,64}", username):
        return jsonify(
            {
                "error": "validation_error",
                "message": "工号仅支持 3-64 位字母、数字、点、下划线和连字符",
            }
        ), 400
    if User.query.filter_by(username=username).first():
        return jsonify(
            {"error": "conflict", "message": "该工号已存在"}
        ), 409
    try:
        _validate_password(password)
        values = _user_payload(data)
    except ValueError as exc:
        return jsonify({"error": "validation_error", "message": str(exc)}), 400
    user = User(username=username, **values)
    user.set_password(password)
    db.session.add(user)
    db.session.flush()
    audit("user.create", "user", user.id, values | {"username": username})
    db.session.commit()
    return jsonify({"user": _serialize_user(user)}), 201


@bp.put("/admin/users/<int:user_id>")
@roles_required("admin")
def update_user(user_id):
    user = db.get_or_404(User, user_id)
    data = request.get_json(silent=True) or {}
    try:
        values = _user_payload(data, user)
    except ValueError as exc:
        return jsonify({"error": "validation_error", "message": str(exc)}), 400
    if user.id == g.current_user.id and (
        values["role"] != "admin" or not values["is_active"]
    ):
        return jsonify(
            {"error": "invalid_state", "message": "不能停用或降级当前登录管理员"}
        ), 409
    if (
        user.role == "admin"
        and user.is_active
        and (values["role"] != "admin" or not values["is_active"])
        and _active_admin_count() <= 1
    ):
        return jsonify(
            {"error": "invalid_state", "message": "系统必须保留至少一个启用管理员"}
        ), 409
    before = _serialize_user(user)
    for field, value in values.items():
        setattr(user, field, value)
    if values["is_active"]:
        user.failed_login_count = 0
        user.locked_until = None
    audit(
        "user.update",
        "user",
        user.id,
        {"before": before, "after": values},
    )
    db.session.commit()
    return {"user": _serialize_user(user, len(user.tasks))}


@bp.post("/admin/users/<int:user_id>/reset-password")
@roles_required("admin")
def reset_user_password(user_id):
    user = db.get_or_404(User, user_id)
    password = str((request.get_json(silent=True) or {}).get("password", ""))
    try:
        _validate_password(password)
    except ValueError as exc:
        return jsonify({"error": "validation_error", "message": str(exc)}), 400
    user.set_password(password)
    user.failed_login_count = 0
    user.locked_until = None
    audit("user.password_reset", "user", user.id)
    db.session.commit()
    return {"message": "密码已重置"}


@bp.post("/admin/users/<int:user_id>/unlock")
@roles_required("admin")
def unlock_user(user_id):
    user = db.get_or_404(User, user_id)
    user.failed_login_count = 0
    user.locked_until = None
    audit("user.unlock", "user", user.id)
    db.session.commit()
    return {"user": _serialize_user(user, len(user.tasks))}


@bp.get("/admin/stats")
@roles_required("admin")
def stats():
    total_tasks = Task.query.filter(Task.deleted_at.is_(None)).count()
    completed = Task.query.filter_by(status="completed").count()
    failed = Task.query.filter_by(status="failed").count()
    total_duration = (
        db.session.query(func.coalesce(func.sum(Task.audio_duration), 0)).scalar()
    )
    return {
        "total_tasks": total_tasks,
        "completed_tasks": completed,
        "failed_tasks": failed,
        "total_audio_duration": total_duration,
        "active_users": User.query.filter_by(is_active=True).count(),
    }
